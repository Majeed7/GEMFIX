import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.datasets import make_classification

import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.datasets import make_classification
from scipy.stats import kendalltau
from sklearn.datasets import load_iris, load_breast_cancer, load_diabetes
import random

from lime.lime_tabular import LimeTabularExplainer
import shap
import matplotlib.pyplot as plt
from scipy.stats import kendalltau, spearmanr
import warnings
warnings.filterwarnings("ignore")

from tabular_datasets import * 
from explainer.gemfix import GEMFIX

from explainer.bishapley_kernel import Bivariate_KernelExplainer

import time 
import openpyxl
from openpyxl import load_workbook

from pathlib import Path

def feature_removing_effect(feature_importance, X_tbx, X_bg, remove_feature):
    
    sorted_features = np.argsort(np.abs(feature_importance), axis=1)
    all_predic_diff = []
    y_x = exp_func(X_tbx)
    for i, x in enumerate(X_tbx):
        X_inverted = np.tile(x, (X_bg.shape[0],1))
        y_x_i = y_x[i]
        predic_diff = []
        
        for j in range(remove_feature):
            X_inverted[:, sorted_features[i,:j]] = X_bg[:,sorted_features[i,:j]]
            y_hat = np.mean(exp_func(X_inverted))             

            predic_diff.append(np.abs(y_x_i - y_hat))
        
        all_predic_diff.append(predic_diff)

    return np.array(all_predic_diff)

def classification_predict(X):
    return model.predict_proba(X)[:,1]

excel_path_results = Path(f'tabular_exp.xlsx')
excel_path_feature_removal = Path(f'tabular_feature_removal.xlsx')
#miles_per_gallon(), stackloos()
datasets = [diabetes(), california_housing(), extramarital_affairs(), mode_choice(),  statlog_heart(), credit_approval(), heart_mortality()]  
sampleNo_tbx = 50

for data in datasets:
    
    # Loading data
    X, y, db_name, mode = data
    print(db_name)

    # Split the data into training and test sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    n_train, d = X_train.shape

    # train a glass box model 
    #model = ExplainableBoostingClassifier() if mode =='classification' else ExplainableBoostingRegressor()
    model = RandomForestClassifier(n_estimators=500) if mode =='classification' else RandomForestRegressor(n_estimators=500)
    model.fit(X_train, y_train)
    y_pred = model.predict(X_test)
    exp_func = classification_predict if mode == 'classification' else model.predict

    if mode == 'regression':
        from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
        # Calculate MAE
        mae = mean_absolute_error(y_test, y_pred)
        print(f"Mean Absolute Error (MAE): {mae:.2f}")

        # Calculate MSE
        mse = mean_squared_error(y_test, y_pred)
        print(f"Mean Squared Error (MSE): {mse:.2f}")

        # Calculate RMSE
        rmse = np.sqrt(mse)  # Or directly use mean_squared_error with squared=False
        print(f"Root Mean Squared Error (RMSE): {rmse:.2f}")

        # Calculate R-squared
        r2 = r2_score(y_test, y_pred)
        print(f"R-squared: {r2:.2f}")
    else: 
        from sklearn.metrics import accuracy_score, classification_report, confusion_matrix, roc_auc_score, roc_curve, log_loss
        y_probs = model.predict_proba(X_test)[:, 1]  # Probabilities for the positive class
        print("Accuracy:", accuracy_score(y_test, y_pred))
        print("Confusion Matrix:\n", confusion_matrix(y_test, y_pred))
        print("Classification Report:\n", classification_report(y_test, y_pred))


    # Get baseline importance
    X_bg = X_train if X_train.shape[0] < 100 else shap.sample(X_train, 100)
    indices = np.random.choice(X_test.shape[0], size=min(sampleNo_tbx, X_test.shape[0]), replace=False)
    X_tbx = X_test[indices,:]

    kshap = shap.KernelExplainer(exp_func, X_bg)
    baseline_importances = kshap.shap_values(X_tbx, nsamples = 2 ** d - 2)

    # SHAP, Bivairate SHAP, and gemfix setup
    shap_explainer = shap.KernelExplainer(exp_func, X_bg)
    shap_sampling_explainer = shap.SamplingExplainer(exp_func, X_bg)
    gemfix_explainer = GEMFIX(exp_func, X_bg)
    bshap = Bivariate_KernelExplainer(exp_func, X_bg)

    # Different numbers of samples for explanation
    max_samplesize = 5000
    sample_no = d #if d < 14 else d*2
    sample_size = random.sample(range(2*d, np.min((2**d, max_samplesize))), sample_no)
    sample_sizes = np.sort(sample_size)

    kshap_diff = []
    kshap_diff_std = []
    kshap_time = []

    sshap_diff = []
    sshap_diff_std = []
    sshap_time = []

    bishap_diff = []
    bishap_diff_std = []
    bishap_time = []

    gemfix_1_diff = []
    gemfix_1_diff_std = []
    gemfix_1_time = []
    
    gemfix_01_diff = []
    gemfix_01_diff_std = []
    gemfix_01_time = []

    gemfix_001_diff = []
    gemfix_001_diff_std = []
    gemfix_001_time = []
    
    
    # Perform experiment
    for num_samples in sample_sizes:
        # Kernel SHAP
        st = time.time()
        shap_values = shap_explainer.shap_values(X_tbx, nsamples=num_samples)        
        endt = time.time() - st 
        diff = np.linalg.norm(shap_values - baseline_importances, 1, axis=1)

        kshap_diff.append(np.mean(diff))    
        kshap_diff_std.append(np.std(diff)) 
        kshap_time.append(endt)             

        
        # Sampling SHAP
        st = time.time()
        shap_sampling_values = shap_sampling_explainer.shap_values(X_tbx, nsamples=num_samples)
        endt = time.time() - st
        diff = np.linalg.norm(shap_sampling_values - baseline_importances, 1, axis=1)

        sshap_diff.append(np.mean(diff))   
        sshap_diff_std.append(np.std(diff))
        sshap_time.append(endt)            

        # Bivariate SHAP
        st = time.time()
        bshap_values = bshap.shap_values(X_tbx, nsamples=num_samples)
        endt = time.time() - st
        diff = np.linalg.norm(bshap_values - baseline_importances, 1, axis=1) # kendalltau(baseline_ranking, shap_rank)[0] #    

        bishap_diff.append(np.mean(diff))    
        bishap_diff_std.append(np.std(diff)) 
        bishap_time.append(endt)             

        
        ## GEMFIX - lam 0.1
        st = time.time()
        gemfix_values = gemfix_explainer.shap_values(X_tbx, nsamples=num_samples, lam=.1)
        endt = time.time() - st
        diff = np.linalg.norm(gemfix_values - baseline_importances, 1, axis=1) # kendalltau(baseline_ranking, gemfix_rank)[0] #    
        
        gemfix_1_diff.append(np.mean(diff))    
        gemfix_1_diff_std.append(np.std(diff)) 
        gemfix_1_time.append(endt)             

        ## GEMFIX - lam 0.01
        st = time.time()
        gemfix_values = gemfix_explainer.shap_values(X_tbx, nsamples=num_samples, lam=.01)
        endt = time.time() - st
        diff = np.linalg.norm(gemfix_values - baseline_importances, 1, axis=1) # kendalltau(baseline_ranking, gemfix_rank)[0] #    
        
        gemfix_01_diff.append(np.mean(diff))    
        gemfix_01_diff_std.append(np.std(diff)) 
        gemfix_01_time.append(endt)             

        ## GEMFIX - lam 0.001
        st = time.time()
        gemfix_values = gemfix_explainer.shap_values(X_tbx, nsamples=num_samples, lam=.001)
        endt = time.time() - st
        diff = np.linalg.norm(gemfix_values - baseline_importances, 1, axis=1) # kendalltau(baseline_ranking, gemfix_rank)[0] #    
        
        gemfix_001_diff.append(np.mean(diff))    
        gemfix_001_diff_std.append(np.std(diff)) 
        gemfix_001_time.append(endt)             


        ## Get the effect of feature removal
        n_sampleNo = len(sample_sizes)
        mid_index = n_sampleNo // 2
        
        if n_sampleNo % 2 == 1:
            mid_value = sample_sizes[mid_index]  # Odd length, return middle item
        else:
            mid_value = sample_sizes[mid_index - 1]
            
        if num_samples == int(mid_value):
            remove_feature = int(np.ceil(d * 0.6))
            
            ## Other methods

            explainer = LimeTabularExplainer(training_data = X_bg, mode = 'regression')
            lime_values = []
            for x in X_tbx:
                explanation = explainer.explain_instance(data_row = x, predict_fn = exp_func, num_features = d)
                exp = [0] * d
                for feature_idx, contribution in explanation.local_exp[0]:
                    exp[feature_idx] = contribution
                lime_values.append(exp)
            
            lime_removal_effect = feature_removing_effect(lime_values, X_tbx, X_bg, remove_feature)

            shap_removal_effect = feature_removing_effect(shap_values, X_tbx, X_bg, remove_feature)
            sshap_removal_effect = feature_removing_effect(shap_values, X_tbx, X_bg, remove_feature)
            bishap_removal_effect = feature_removing_effect(shap_values, X_tbx, X_bg, remove_feature)
            gemfix_1_removal_effect = feature_removing_effect(shap_values, X_tbx, X_bg, remove_feature)
            gemfix_01_removal_effect = feature_removing_effect(shap_values, X_tbx, X_bg, remove_feature)
            gemfix_001_removal_effect = feature_removing_effect(shap_values, X_tbx, X_bg, remove_feature)


    ## Storing the difference and execution time of the methods
    row_header = ['Kernel SHAP', 'Kernel SHAP std', 'Kernel SHAP time', 'Sampling SHAP', 'Sampling SHAP std', 'Sampling SHAP time',
                  'Bivariate SHAP', 'Bivariate SHAP std', 'Bivariate SHAP time', 'GEMFIX_0.1', 'GEMFIX_0.1 std', 'GEMFIX_0.1 time',
                  'GEMFIX_0.01', 'GEMFIX_0.01 std', 'GEMFIX_0.01 time', 'GEMFIX_0.001', 'GEMFIX_0.001 std', 'GEMFIX_0.001 time']
    
    colum_header = sample_sizes

    all_results = [kshap_diff, kshap_diff_std, kshap_time, sshap_diff, sshap_diff_std, sshap_time, bishap_diff, bishap_diff_std, bishap_time,
                   gemfix_1_diff, gemfix_1_diff_std, gemfix_1_time, gemfix_01_diff, gemfix_01_diff_std, gemfix_01_time, gemfix_001_diff, gemfix_001_diff_std, gemfix_001_time] 
    
    df = pd.DataFrame(all_results, index=row_header, columns=sample_sizes)

    mode = 'a' if excel_path_results.exists() else 'w'
    
    with pd.ExcelWriter(excel_path_results, engine='openpyxl', mode=mode) as writer:
        # Attempt to load the workbook if it exists to check for sheet names
        if mode == 'a':
            writer_book = load_workbook(excel_path_results)
            writer_sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        sheet_name = db_name

        counter = 1
        while sheet_name in writer_sheets:
            sheet_name = f"{sheet_name}_{(counter)}"
            counter += 1
        
        df.to_excel(writer, sheet_name=sheet_name, index_label='Method')

    ## Saving the feature removal effect
    row_header = ['Kernel SHAP', 'Kernel SHAP std', 'Sampling SHAP', 'Sampling SHAP std', 'Bivariate SHAP', 'Bivariate SHAP std',
                  'GEMFIX_0.1', 'GEMFIX_0.1 std', 'GEMFIX_0.01', 'GEMFIX_0.01 std', 'GEMFIX_0.001', 'GEMFIX_0.001 std']
    
    column_header = np.arange(1, remove_feature+1)

    all_results = [np.mean(shap_removal_effect, axis=0), np.std(shap_removal_effect, axis=0), np.mean(sshap_removal_effect, axis=0), np.std(sshap_removal_effect, axis=0),
                   np.mean(bishap_removal_effect, axis=0), np.std(bishap_removal_effect, axis=0), np.mean(gemfix_1_removal_effect, axis=0), np.std(gemfix_1_removal_effect, axis=0),
                   np.mean(gemfix_01_removal_effect, axis=0), np.std(gemfix_01_removal_effect, axis=0), np.mean(gemfix_001_removal_effect, axis=0), np.std(gemfix_001_removal_effect, axis=0)] 
    
    df = pd.DataFrame(all_results, index=row_header, columns=column_header)

    mode = 'a' if excel_path_feature_removal.exists() else 'w'
    
    with pd.ExcelWriter(excel_path_feature_removal, engine='openpyxl', mode=mode) as writer:
        # Attempt to load the workbook if it exists to check for sheet names
        if mode == 'a':
            writer_book = load_workbook(excel_path_results)
            writer_sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        sheet_name = db_name

        counter = 1
        while sheet_name in writer.sheets:
            sheet_name = f"{sheet_name}_{(counter)}"
            counter += 1
        
        df.to_excel(writer, sheet_name=sheet_name, index_label='Method')

print("done!")