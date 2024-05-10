import lime.lime_text
import transformers
from transformers import BertTokenizerFast
import numpy as np
import time
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from pathlib import Path 
from datasets import load_dataset
import random 

import shap
from explainer.gemfix import GEMFIX
from explainer.bishapley_kernel import Bivariate_KernelExplainer
import lime
import lime.lime_tabular

model = transformers.pipeline('sentiment-analysis')
tokenizer = BertTokenizerFast.from_pretrained('bert-base-cased')

results_xsl = Path(f'text_exp.xlsx')

def encode(batch, max_length = 400, tokenizer = BertTokenizerFast.from_pretrained('bert-base-cased'), add_special_tokens = False):
    return tokenizer(batch, padding="longest", truncation=True, max_length=max_length, add_special_tokens = add_special_tokens)['input_ids']

def sentiment_probability(text_list):
    probs = []
    outputs = model(text_list)
    for ot in outputs:
        value = ot['score']
        if ot['label'].lower() == 'negative': value *= -1
        prob = .5 * (value + 1)
        probs.append([1 - prob, prob])
    
    return np.array(probs)

def adjust_sheet_name(existing_sheets, desired_name):
    """ Generate a new sheet name if the desired one already exists """
    count = 1
    new_name = desired_name
    while new_name in existing_sheets:
        new_name = f"{desired_name} ({count})"
        count += 1
    return new_name
# define utility function wrapper for the huggingface model
# utility function takes input as x, and returns a score
class eval_transformer():
    def __init__(self, model, tokenizer = BertTokenizerFast.from_pretrained('bert-base-cased'), **kwargs):
        self.model = model
        #self.baseline = None
        self.tokenizer = tokenizer

    def __call__(self, x, **kwargs):   
        n = len(x) #x.shape[0]
        x = x.astype(dtype = 'int')
        #x = x.tolist()
        x_tkn = self.tokenizer.batch_decode(x)
        output = self.model(x_tkn)
        score = []
        for oput in output:
            sign = -1 if oput['label'].lower() == 'negative' else 1
            score.append(sign * oput['score'])
        
        return np.array(score)

def feature_removing_effect(feature_importance, x, model, baseline, remove_feature_no):
    
    sorted_features = np.argsort(np.abs(feature_importance)).squeeze()
    y_x = model(x)
    x_inverted = x.copy().squeeze()
    predic_diff = []
    
    for j in range(remove_feature_no):
        x_inverted[sorted_features[:j+1]] = baseline
        y_hat = model(x_inverted.reshape(1,-1))             

        predic_diff.append(np.abs(y_x.squeeze() - y_hat.squeeze()))
    

    return predic_diff

def shap_removal_effect(x_tkn, value_function, baseline_char, remove_feature_no, nsamples=100):
    x_train = np.zeros_like(x_tkn) + baseline_char
    explainer = shap.KernelExplainer(value_function, x_train, nsamples = nsamples)
    phi = explainer.shap_values(x_tkn, nsamples = nsamples)
    
    predic_diff = feature_removing_effect(phi, x_tkn, value_function, baseline_char, remove_feature_no)

    return predic_diff, phi

def bishap_removal_effect(x_tkn, value_function, baseline_char, remove_feature_no, nsamples=100):
    x_train = np.zeros_like(x_tkn) + baseline_char
    explainer = Bivariate_KernelExplainer(value_function, x_train, nsamples = nsamples)
    phi = explainer.shap_values(x_tkn, nsamples = nsamples)
    
    predic_diff = feature_removing_effect(phi, x_tkn, value_function, baseline_char, remove_feature_no)

    return predic_diff, phi

def sshap_removal_effect(x_tkn, value_function, baseline_char, remove_feature_no, nsamples=100):
    x_train = np.zeros_like(x_tkn) + baseline_char
    explainer = shap.SamplingExplainer(value_function, x_train, nsamples = nsamples)
    phi = explainer.shap_values(x_tkn, l1_reg = False, nsamples=nsamples)
    
    predic_diff = feature_removing_effect(phi, x_tkn, value_function, baseline_char, remove_feature_no)

    return predic_diff, phi

def gemfix_removal_effect(x_tkn, value_function, baseline_char, remove_feature_no, nsamples=100, lam=0.001):
    x_train = np.zeros_like(x_tkn) + baseline_char
    explainer = GEMFIX(value_function, x_train, nsamples = nsamples)
    phi = explainer.shap_values(x_tkn, nsamples = nsamples, lam=lam)
    
    predic_diff = feature_removing_effect(phi, x_tkn, value_function, baseline_char, remove_feature_no)

    return predic_diff, phi

def lime_removal_effect(x_text, x_tkn, value_function, sentiment_probability, remove_feature_no, baseline_char, nsamples=100):
    lime_txt = lime.lime_text.LimeTextExplainer(class_names = ["negative", "positive"])
    lime_exp = lime_txt.explain_instance(x_text, sentiment_probability, num_samples=nsamples)
    phi = np.zeros((len(x_tkn[0]), ))
    for fs in lime_exp.local_exp[1]:
        phi[fs[0]] = fs[1]

    predic_diff = feature_removing_effect(phi, x_tkn, value_function, baseline_char, remove_feature_no)

    return predic_diff, phi

if __name__ == '__main__':

    value_function = eval_transformer(model)
    
    # sstds = load_dataset("glue", "sst2")
    # test_data = sstds['test']['sentence']
    

    # dataset = load_dataset("imdb")
    # test_data = dataset['test']['text']

    # dataset = load_dataset("sentiment140")
    # test_data = dataset['test']['text']

    # dataset = load_dataset("yelp_review_full")
    # test_data = dataset['test']['text']

    # dataset = load_dataset("amazon_polarity")
    # test_data = dataset['test']['content']

    # dataset = load_dataset("rotten_tomatoes")
    # test_data = dataset['test']['text']

    dataset = load_dataset("go_emotions")
    test_data = dataset['test']['text']

    tokenized_train = tokenizer(test_data, truncation = True).data['input_ids']
    sst_len = [len(i) for i in tokenized_train]
    selected_review = np.where( (np.array(sst_len) <= 35) & (np.array(sst_len) >= 25) )[0]


    review_tbx = 50
    if len(selected_review) < review_tbx: 
        review_samples = selected_review
    else:
        review_samples = random.sample(list(selected_review), review_tbx)

    small_test_dataset = [test_data[i] for i in review_samples]
    tokenized_train = [tokenized_train[i] for i in review_samples]


    phis = []
    baseline_char = 103
    all_feature_no = [] 

    shap_predic_diff = []
    gemfix_predic_diff = []
    gemfix01_predic_diff = []
    gemfix001_predic_diff = []
    bishap_predic_diff = []
    sshap_predic_diff = []
    lime_predic_diff = []

    shap_time = []
    gemfix_time = []    
    bishap_time = []
    sshap_time = []
    lime_time = []
    nsamples = 500
    for i in range(len(tokenized_train)):
        print(f"{i}")
        x_text = small_test_dataset[i]
        x_tkn = np.array(tokenized_train[i]).reshape(1,-1)

        remove_feature_no = int(np.floor( len(x_tkn[0]) * 0.95 ))
        all_feature_no.append(len(x_tkn[0]))
        
        stime = time.time()
        predict_diff, phi = shap_removal_effect(x_tkn, value_function, baseline_char, remove_feature_no, nsamples=nsamples)
        shap_predic_diff.append(predict_diff)
        shap_time.append(time.time() - stime)

        stime = time.time()
        predict_diff, phi = gemfix_removal_effect(x_tkn, value_function, baseline_char, remove_feature_no, nsamples=nsamples, lam=.001)
        gemfix_predic_diff.append(predict_diff)
        gemfix_time.append(time.time() - stime)

        predict_diff, phi = gemfix_removal_effect(x_tkn, value_function, baseline_char, remove_feature_no, nsamples=nsamples, lam=0.01)
        gemfix01_predic_diff.append(predict_diff)

        predict_diff, phi = gemfix_removal_effect(x_tkn, value_function, baseline_char, remove_feature_no, nsamples=nsamples, lam=0.001)
        gemfix001_predic_diff.append(predict_diff)


        stime = time.time()
        predict_diff, phi = sshap_removal_effect(x_tkn, value_function, baseline_char, remove_feature_no, nsamples=nsamples)
        sshap_predic_diff.append(predict_diff)
        sshap_time.append(time.time() - stime)

        stime = time.time()
        predict_diff, phi = bishap_removal_effect(x_tkn, value_function, baseline_char, remove_feature_no, nsamples=nsamples)
        bishap_predic_diff.append(predict_diff)
        bishap_time.append(time.time() - stime)

        stime = time.time()
        predict_diff, phi = lime_removal_effect(x_text, x_tkn, value_function, sentiment_probability, remove_feature_no, baseline_char, nsamples=100)
        lime_predic_diff.append(predict_diff)
        lime_time.append(time.time() - stime)

    method_names = ['Kernel SHAP', 'Sampling SHAP', 'Bivariate SHAP', 'LIME', 'GEMFIX', 'GEMFIX_01', 'GEMFIX_001']
    all_resutlts = [(shap_time, shap_predic_diff), (sshap_time, sshap_predic_diff), (bishap_time, bishap_predic_diff), (lime_time, lime_predic_diff), 
                    (gemfix_time, gemfix_predic_diff), (gemfix_time, gemfix01_predic_diff), (gemfix_time, gemfix001_predic_diff)]
    mode = 'a' if results_xsl.exists() else 'w'
    with pd.ExcelWriter(results_xsl, engine='openpyxl', mode=mode) as writer:
        if mode == 'a':
            # Load the existing workbook to check sheet names
            writer.book = load_workbook(results_xsl)
            existing_sheets = writer.book.sheetnames
        else:
            existing_sheets = []

        for (time, list_), name in zip(all_resutlts, method_names):
            # Adjust the sheet name with current time
            adjusted_name = adjust_sheet_name(existing_sheets, name)
            
            # Convert each list to a DataFrame
            df = pd.DataFrame(list_)
            df['time'] = time
            df['all_feature_no'] = all_feature_no
            
            # Write each DataFrame to a specific sheet
            df.to_excel(writer, sheet_name=adjusted_name, index=False, header=False)


    print("done!")