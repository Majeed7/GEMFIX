import numpy as np
import matplotlib.pyplot as plt 

import shap 
from explainer.gemfix import GEMFIX
from explainer.bishapley_kernel import Bivariate_KernelExplainer
from shapreg import removal, games, shapley
from explainer.MAPLE import MAPLE
from lime import lime_tabular

from pathlib import Path
import pandas as pd 
from openpyxl import load_workbook

results_xsl = Path('synthesized_results.xlsx')

def create_rank(scores): 
	"""
	Compute rank of each feature based on weight.
	
	"""
	scores = abs(scores)
	n, d = scores.shape
	ranks = []
	for i, score in enumerate(scores):
		# Random permutation to avoid bias due to equal weights.
		idx = np.random.permutation(d) 
		permutated_weights = score[idx]  
		permutated_rank=(-permutated_weights).argsort().argsort()+1
		rank = permutated_rank[np.argsort(idx)]

		ranks.append(rank)

	return np.array(ranks)

def generate_X(n_samples=100, n_features=10):
    #return np.random.uniform(-1, 1, (n_samples, n_features))
    return np.random.randn(n_samples, n_features)

def generate_dataset_poly_sine(n_samples=100, n_features=10):
    X = generate_X(n_samples, n_features)
    
    def fn(X):
        f1, f2, f3 = X[:, 0], X[:, 1], X[:, 2]
        y = f1**2 - 0.5 * f2**2 + f3**3 + np.sin(2 * np.pi * f1) 

        logit = np.exp(y) 
        prob_1 = np.expand_dims(1 / (1+logit) ,1)

        return prob_1
    
    return X, fn(X), fn, np.arange(0,3), 'Poly Sine'

def generate_dataset_squared_exponentials(n_samples=100, n_features=10):
    X = generate_X(n_samples, n_features)
    
    def fn(X):
        logit = np.exp(np.sum(X[:,:4]**2, axis = 1) - 4.0) 

        prob_1 = np.expand_dims(1 / (1+logit) ,1)
        prob_0 = np.expand_dims(logit / (1+logit) ,1)

        #y = np.concatenate((prob_0,prob_1), axis = 1)

        return prob_1
    
    return X, fn(X), fn, np.arange(0,4), 'squared Expoenetial'

def generate_additive_labels(n_samples=100, n_features=10):
        X = generate_X(n_samples, n_features)

        def fn(X):
            logit = np.exp(-100 * np.sin(0.2*X[:,0]) + abs(X[:,1]) + X[:,2]) + np.exp(-X[:,3])  

            prob_1 = np.expand_dims(1 / (1+logit) ,1)
            prob_0 = np.expand_dims(logit / (1+logit) ,1)

            y = np.concatenate((prob_0,prob_1), axis = 1)
            return prob_1

        return X, fn(X), fn, np.arange(0,4), "Nonlinear Additive"

def generate_XOR(n_samples=100, n_features=10):
    X = generate_X(n_samples, n_features)
    
    def fn(X):
        y = 0.5 * ( np.exp(X[:,0]*X[:,1]*X[:,2]) + np.exp(X[:,3]*X[:,4]))
        prob_1 = np.expand_dims(1 / (1+y) ,1)

        return y

    return X, fn(X), fn, np.arange(0,5), "XOR data set"

def generate_simple_interactions(n_samples=100, n_features=10):
    X = generate_X(n_samples, n_features)
    
    def fn(X):
        y = ((X[:,0]*X[:,1]*X[:,2]) + (X[:,3]*X[:,4]))

        return y

    return X, fn(X), fn, np.arange(0,5), "XOR data set"

if __name__ == '__main__':
    np.random.seed(10)

    X_sample_no = 100  # number of sampels for generating explanation
    smaple_tbX = 100   # number of samples to be explained
    sample_no_gn = 100 # number of generated synthesized instances 
    feature_no_gn = 8 # number of features for the synthesized instances

    # Example usage of one of the functions
    X, y, fn, feature_imp, ds_name = generate_simple_interactions(sample_no_gn, feature_no_gn)
    
    ## GEMFIX
    gemfix = GEMFIX(fn, X, lam=0.001)
    gem_values = gemfix.shap_values(X, nsamples=X_sample_no)
    gem_ranks = create_rank(np.array(gem_values).squeeze())
    gem_avg_ranks = np.mean(gem_ranks[:,feature_imp], axis=1)
    gemfix_mean_rank = np.mean(gem_avg_ranks)

    ## SHAP
    explainer = shap.KernelExplainer(fn, X, l1_reg=False)
    shap_values = explainer.shap_values(X, nsamples=X_sample_no, l1_reg=False)
    shap_ranks = create_rank(shap_values.squeeze())
    shap_avg_ranks = np.mean(shap_ranks[:,feature_imp], axis=1)
    shap_mean_rank = np.mean(shap_avg_ranks)

    ## Sampling SHAP
    sexplainer = shap.SamplingExplainer(fn, X, l1_reg=False)
    sshap_values = sexplainer.shap_values(X, nsamples=X_sample_no, l1_reg=False, min_samples_per_feature=1)
    sshap_ranks = create_rank(sshap_values.squeeze())
    sshap_avg_ranks = np.mean(sshap_ranks[:,feature_imp], axis=1)
    sshap_mean_rank = np.mean(sshap_avg_ranks)


    plt.boxplot([gem_avg_ranks, shap_avg_ranks, sshap_avg_ranks])
    ## Bivariate SHAP
    bishap = Bivariate_KernelExplainer(fn, X)
    bishap_values = bishap.shap_values(X, nsamples=X_sample_no, l1_reg=False)
    bishap_ranks = create_rank(np.array(bishap_values).squeeze())
    bishap_avg_ranks = np.mean(bishap_ranks[:,feature_imp], axis=1)
    bishap_mean_rank = np.mean(bishap_avg_ranks)


    ## LIME, Unbiased SHAP, and MAPLE 
    lime_exp = lime_tabular.LimeTabularExplainer(X, discretize_continuous=False, mode="regression")
    imputer = removal.MarginalExtension(X, fn)
    exp_maple = MAPLE(X, y, X, y)

    ushap_values = np.empty_like(X)
    lime_values = np.empty_like(X)
    maple_values = np.empty_like(X)
    for i in range(X.shape[0]):
        x = X[i, ]
    
        ## Unbiased kernel shap 
        game = games.PredictionGame(imputer, x)
        values = shapley.ShapleyRegression(game, n_samples=X_sample_no, paired_sampling=False)
        ushap_values[i,:] = values.values.squeeze()

        ## LIME 
        exp = lime_exp.explain_instance(x, fn, num_samples = X_sample_no)
            
        for tpl in exp.as_list():
            lime_values[i, int(tpl[0])] = tpl[1]

        ## MAPLE
        mpl_exp = exp_maple.explain(x)
        maple_values[i,] = (mpl_exp['coefs'][1:]).squeeze()


    lime_ranks = create_rank(lime_values)
    lime_avg_ranks = np.mean(lime_ranks[:,feature_imp], axis=1)
    lime_mean_rank = np.mean(lime_avg_ranks)

    maple_ranks = create_rank(maple_values)
    maple_avg_ranks = np.mean(maple_ranks[:,feature_imp], axis=1)
    maple_mean_rank = np.mean(maple_avg_ranks)

    ushap_ranks = create_rank(ushap_values)
    ushap_avg_ranks = np.mean(ushap_ranks[:,feature_imp], axis=1)
    ushap_mean_rank = np.mean(ushap_avg_ranks)

    plt.boxplot([gem_avg_ranks, shap_avg_ranks, bishap_avg_ranks, sshap_avg_ranks, ushap_avg_ranks, lime_avg_ranks, maple_avg_ranks])


    method_names = ['GEM-FIX', 'Kernel SHAP', 'Sampling SHAP', 'Unbiased SHAP', 'Bivariate SHAP', 'LIME',  'MAPLE']
    all_results = [gem_avg_ranks, shap_avg_ranks, sshap_avg_ranks, ushap_avg_ranks, bishap_avg_ranks, lime_avg_ranks, maple_avg_ranks]

    df = pd.DataFrame(all_results, index=method_names)

    mode = 'a' if results_xsl.exists() else 'w'
    with pd.ExcelWriter(results_xsl, engine='openpyxl', mode=mode) as writer:
        # if mode == 'a':
        #     # Load the existing workbook to check sheet names
        #     writer.book = load_workbook(results_xsl)
        #     existing_sheets = writer.book.sheetnames
        # else:
        #     existing_sheets = []
    
        # Write each DataFrame to a specific sheet
        df.to_excel(writer, sheet_name=ds_name, index_label='Method')

    print("done!")
    

